from __future__ import annotations
import openpyxl
from collections import defaultdict
from typing import Dict, Tuple, List, Any, Optional
import os

REQUIRED_COLUMNS = [
    "Type (Monoaxial/Polyaxial/Cannulated/Uniaxial)",
    "Diameter (mm)",
    "Length (mm)",
    "Product #",
    "Stock Location",
    "Pan #",
    "Quantity in Pan",
    "Number of Pans",
    "Total in Hospital",
]

def get_default_inventory_path() -> str:
    # inventory_loader.py lives in inventory/
    base = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(base, "Inventory.xlsx")  # adjust if you store elsewhere

def load_inventory_excel(filepath: str, sheet_name: Optional[str] = None):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active

    header_row_idx, col_map = _find_header_row_and_map(ws)
    _validate_required_columns(col_map)

    totals_by_key: Dict[Tuple[str, float, int], int] = defaultdict(int)
    rows_by_key: Dict[Tuple[str, float, int], List[Dict[str, Any]]] = defaultdict(list)

    # Iterate through data rows
    for r in range(header_row_idx + 1, ws.max_row + 1):
        row_values = {}
        empty_count = 0

        for col_name, c in col_map.items():
            val = ws.cell(row=r, column=c).value
            row_values[col_name] = val
            if val is None or (isinstance(val, str) and val.strip() == ""):
                empty_count += 1

        if empty_count >= len(col_map) - 1:
            continue

        type_str = _clean_type(row_values.get(REQUIRED_COLUMNS[0]))
        diameter = _to_float(row_values.get("Diameter (mm)"))
        length = _to_int(row_values.get("Length (mm)"))

        if not type_str or diameter is None or length is None:
            continue

        key = (type_str, diameter, length)

        product_no = _to_str(row_values.get("Product #"))
        stock_loc = _to_str(row_values.get("Stock Location"))
        pan = _to_str(row_values.get("Pan #"))
        qty_in_pan = _to_int(row_values.get("Quantity in Pan"))
        num_pans = _to_int(row_values.get("Number of Pans"))
        total_in_hosp = _to_int(row_values.get("Total in Hospital"))

        rows_by_key[key].append(
            {
                "type": type_str,
                "diameter_mm": diameter,
                "length_mm": length,
                "product_number": product_no,
                "stock_location": stock_loc,
                "pan": pan,
                "quantity_in_pan": qty_in_pan,
                "number_of_pans": num_pans,
                "total_in_hospital_cell": total_in_hosp,
            }
        )

        if total_in_hosp is not None:
            totals_by_key[key] += total_in_hosp

    # Fallback totals for keys where Total in Hospital was never provided
    for key, rows in rows_by_key.items():
        if totals_by_key.get(key, 0) > 0:
            continue

        fallback_total = 0
        for row in rows:
            q = row.get("quantity_in_pan")
            n = row.get("number_of_pans")
            if q is None or n is None:
                continue
            fallback_total += q * n

        totals_by_key[key] = fallback_total

    # Convert defaultdict to normal dicts
    return dict(totals_by_key), dict(rows_by_key)

# -------------------------
# Helpers
# -------------------------

def _find_header_row_and_map(ws):
    max_scan = min(30, ws.max_row)
    normalized_required = [_norm_header(c) for c in REQUIRED_COLUMNS]

    for r in range(1, max_scan + 1):
        row_norm = []
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            row_norm.append(_norm_header(v))

        # Build map of normalized header -> column index (first occurrence)
        temp = {}
        for idx, name in enumerate(row_norm, start=1):
            if name and name not in temp:
                temp[name] = idx

        if all(req in temp for req in normalized_required):
            col_map = {}
            for original, req_norm in zip(REQUIRED_COLUMNS, normalized_required):
                col_map[original] = temp[req_norm]
            return r, col_map

    raise ValueError("Could not find a header row with the required inventory columns.")


def _validate_required_columns(col_map):
    missing = [c for c in REQUIRED_COLUMNS if c not in col_map]
    if missing:
        raise ValueError(f"Missing required columns in inventory sheet: {missing}")


def _norm_header(x):
    if x is None:
        return ""
    s = str(x).strip().lower()
    # normalize weird spacing
    s = " ".join(s.split())
    return s


def _to_str(x):
    if x is None:
        return ""
    return str(x).strip()


def _clean_type(x):
    s = _to_str(x)
    if not s:
        return ""
    s2 = s.lower()
    if "cann" in s2:
        return "Cannulated"
    if "poly" in s2:
        return "Polyaxial"
    if "mono" in s2:
        return "Monoaxial"
    if "uniax" in s2 or "uni" in s2:
        return "Uniaxial"
    return s


def _to_int(x):
    if x is None:
        return None
    if isinstance(x, bool):
        return None
    if isinstance(x, int):
        return x
    if isinstance(x, float):
        # treat 12.0 as 12
        return int(round(x))
    s = str(x).strip()
    if s == "":
        return None
    try:
        return int(float(s))
    except Exception:
        return None


def _to_float(x):
    if x is None:
        return None
    if isinstance(x, bool):
        return None
    if isinstance(x, (int, float)):
        return float(x)
    s = str(x).strip()
    if s == "":
        return None
    try:
        return float(s)
    except Exception:
        return None
    
def validate_inventory_data(totals_by_key, rows_by_key):
    if not rows_by_key:
        return False, "No valid inventory rows were detected. Check that the sheet contains screw rows under the header.", {}

    # Count total parsed rows
    total_rows = sum(len(v) for v in rows_by_key.values())
    if total_rows < 10:
        return False, f"Only {total_rows} valid rows were detected. This looks too small, are you sure this is the correct file?", {}

    allowed_types = {"Monoaxial", "Polyaxial", "Cannulated", "Uniaxial"}
    bad_type = 0
    bad_dims = 0
    zero_totals = 0

    for key, rows in rows_by_key.items():
        t, d, L = key

        if t not in allowed_types:
            bad_type += len(rows)

        if d is None or d < 2.0 or d > 12.0:
            bad_dims += len(rows)

        if L is None or L < 10 or L > 150:
            bad_dims += len(rows)

        if totals_by_key.get(key, 0) <= 0:
            zero_totals += 1

    if bad_type / max(1, total_rows) > 0.25:
        return False, "Too many rows have unrecognized screw types. Expected Monoaxial, Polyaxial, Cannulated, or Uniaxial.", {}

    if bad_dims / max(1, total_rows) > 0.25:
        return False, "Too many rows have invalid Diameter or Length values. Check units and formatting in the spreadsheet.", {}

    stats = {
        "groups": len(rows_by_key),
        "rows": total_rows,
        "groups_with_zero_total": zero_totals,
    }

    warn = ""
    if zero_totals / max(1, len(rows_by_key)) > 0.5:
        warn = " Note: many groups have zero totals, the file may be missing Quantity in Pan or Number of Pans values."

    return True, f"Inventory looks valid. Parsed {stats['rows']} rows across {stats['groups']} screw groups.{warn}", stats

