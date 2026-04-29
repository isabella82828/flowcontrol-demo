import tkinter as tk
from tkinter import ttk, messagebox
import os
import openpyxl
import re

from .rodfather.bin_overview import compute_bin_overview

WHITE = "#FFFFFF"
LOGO_GREEN = "#036160"
FONT = ("Segoe UI", 12)

# Default predicted lengths shown as placeholders until Brian's software is integrated
DEFAULT_PREDICTED_LEFT_MM = 450.0
DEFAULT_PREDICTED_RIGHT_MM = 460.0


class Page11RodSelection:
    def __init__(self, app):
        self.app = app

        # Rod diameter dropdown vars
        self.left_rod_var = tk.StringVar(value="")
        self.right_rod_var = tk.StringVar(value="")

        # Predicted rod lengths — start empty, populate once both sides are selected
        self.left_predicted_var = tk.StringVar(value="")
        self.right_predicted_var = tk.StringVar(value="")

        # Cross-connector widgets/state
        self.cross_connector_var = tk.BooleanVar(value=False)
        self.cross_options = []
        self.cross_selected = []

        self.other_rod_var = tk.StringVar(value="")

        # UI refs
        self.cross_section_frame = None
        self.cross_checks_frame = None
        self.cross_summary_label = None
        self.cross_check_vars = {}

        self.rod_options = [
            "Cobalt Chrome 6.0 mm",
            "Cobalt Chrome 5.5 mm",
            "Titanium 5.5 mm",
            "Titanium 6.0 mm",
        ]

        # Rod Father Excel path
        self.rodfather_path = os.path.join(
            os.path.dirname(__file__),
            "rodfather",
            "Rod-Father.xlsx"
        )

        # Rod Father data
        self.offcuts = []
        self.offcuts_load_error = None

        # Per-side Rod Father request vars
        # Material is derived from rod selection, not user-editable
        self.left_rf_shape_var = tk.StringVar(value="Hex")
        self.left_rf_length_var = tk.StringVar(value="")

        self.right_rf_shape_var = tk.StringVar(value="Hex")
        self.right_rf_length_var = tk.StringVar(value="")

        # Material display vars (read-only labels, derived from rod selection)
        self.left_rf_material_var = tk.StringVar(value="—")
        self.right_rf_material_var = tk.StringVar(value="—")

        # Selected offcuts for plan
        self.left_offcut_id_var = tk.StringVar(value="")
        self.right_offcut_id_var = tk.StringVar(value="")

        # UI refs
        self.offcuts_tree = None
        self.offcuts_status_label = None
        self.alloc_status_label = None
        self.bin_labels = {}

        # Trace rod selections to keep material labels in sync
        self.left_rod_var.trace_add("write", lambda *_: self._on_rod_selection_changed())
        self.right_rod_var.trace_add("write", lambda *_: self._on_rod_selection_changed())

        # Trace predicted lengths to keep RF length fields in sync
        self.left_predicted_var.trace_add("write", lambda *_: self._sync_rf_lengths())
        self.right_predicted_var.trace_add("write", lambda *_: self._sync_rf_lengths())

    # ─────────────────────────────────────────────────────────────────────────
    # setup
    # ─────────────────────────────────────────────────────────────────────────

    def setup(self):
        scrollable = self.app.create_standard_page(
            title_text="Rod Selection",
            back_command=self.app.setup_page_10,
            next_command=self.app.setup_page_12
        )

        ttk.Label(
            scrollable,
            text="Select left and right rod diameters, and optionally add a cross-connector.",
            font=FONT,
            background=WHITE,
            justify="left"
        ).pack(anchor="w", pady=(6, 10))

        # ── 1. Rod selection ─────────────────────────────────────────────────
        table = ttk.LabelFrame(scrollable, text="Rod Selection")
        table.pack(fill="x", pady=(0, 12), padx=0)

        inner = ttk.Frame(table)
        inner.pack(fill="x", padx=12, pady=12)

        ttk.Label(inner, text="", font=FONT).grid(row=0, column=0, sticky="w")
        ttk.Label(inner, text="Left Side", font=("Segoe UI", 11, "bold")).grid(row=0, column=1, sticky="w", padx=(8, 0))
        ttk.Label(inner, text="Right Side", font=("Segoe UI", 11, "bold")).grid(row=0, column=2, sticky="w", padx=(8, 0))
        ttk.Label(inner, text=" ", font=("Segoe UI", 11, "bold")).grid(row=0, column=3, sticky="w", padx=(8, 0))

        ttk.Label(inner, text="Rod Diameter", font=FONT).grid(row=1, column=0, sticky="w", pady=(8, 0))

        left_combo = ttk.Combobox(
            inner, textvariable=self.left_rod_var,
            values=self.rod_options, state="readonly", width=24
        )
        left_combo.grid(row=1, column=1, sticky="w", padx=(8, 0), pady=(8, 0))
        left_combo.bind("<<ComboboxSelected>>", lambda e: self._persist())

        right_combo = ttk.Combobox(
            inner, textvariable=self.right_rod_var,
            values=self.rod_options, state="readonly", width=24
        )
        right_combo.grid(row=1, column=2, sticky="w", padx=(8, 0), pady=(8, 0))
        right_combo.bind("<<ComboboxSelected>>", lambda e: self._persist())

        other_frame = ttk.Frame(inner)
        other_frame.grid(row=1, column=3, sticky="w", padx=(8, 0), pady=(8, 0))
        ttk.Label(other_frame, text="Other:", font=FONT).pack(side="left")
        other_entry = ttk.Entry(other_frame, textvariable=self.other_rod_var, width=18)
        other_entry.pack(side="left", padx=(6, 0))
        other_entry.bind("<KeyRelease>", lambda e: self._persist())

        inner.grid_columnconfigure(0, weight=0)
        inner.grid_columnconfigure(1, weight=1)
        inner.grid_columnconfigure(2, weight=1)
        inner.grid_columnconfigure(3, weight=1)

        # ── 2. Predicted Rod Lengths ─────────────────────────────────────────
        pred_frame = ttk.LabelFrame(scrollable, text="Predicted Rod Lengths")
        pred_frame.pack(fill="x", pady=(0, 12), padx=0)

        pred_inner = ttk.Frame(pred_frame)
        pred_inner.pack(fill="x", padx=12, pady=12)

        ttk.Label(pred_inner, text="", font=FONT).grid(row=1, column=0, sticky="w")
        ttk.Label(pred_inner, text="Left Side", font=("Segoe UI", 11, "bold")).grid(row=1, column=1, sticky="w", padx=(8, 0))
        ttk.Label(pred_inner, text="Right Side", font=("Segoe UI", 11, "bold")).grid(row=1, column=2, sticky="w", padx=(24, 0))

        ttk.Label(pred_inner, text="Predicted length (mm)", font=FONT).grid(row=2, column=0, sticky="w", pady=(6, 0))

        ttk.Label(
            pred_inner, textvariable=self.left_predicted_var,
            font=("Segoe UI", 11, "bold"), foreground=LOGO_GREEN, width=10, anchor="w",
        ).grid(row=2, column=1, sticky="w", padx=(8, 0), pady=(6, 0))

        ttk.Label(
            pred_inner, textvariable=self.right_predicted_var,
            font=("Segoe UI", 11, "bold"), foreground=LOGO_GREEN, width=10, anchor="w",
        ).grid(row=2, column=2, sticky="w", padx=(24, 0), pady=(6, 0))

        # ── 3. Cross connector ───────────────────────────────────────────────
        cc_frame = ttk.LabelFrame(scrollable, text="Cross Connector")
        cc_frame.pack(fill="x", pady=(0, 10))

        cc_inner = ttk.Frame(cc_frame)
        cc_inner.pack(fill="x", padx=12, pady=12)

        ttk.Checkbutton(
            cc_inner, text="Cross-Connector",
            variable=self.cross_connector_var, command=self._on_cross_toggle
        ).pack(anchor="w")

        self.cross_section_frame = ttk.Frame(cc_inner)
        self.cross_section_frame.pack(fill="x", pady=(10, 0))

        ttk.Label(
            self.cross_section_frame,
            text="Select up to 2 levels for cross-connector placement:",
            font=("Segoe UI", 10)
        ).pack(anchor="w", pady=(0, 6))

        self.cross_checks_frame = ttk.Frame(self.cross_section_frame)
        self.cross_checks_frame.pack(fill="x")

        self.cross_summary_label = ttk.Label(
            self.cross_section_frame, text="Selected: None", font=("Segoe UI", 10)
        )
        self.cross_summary_label.pack(anchor="w", pady=(6, 0))

        self._restore()
        if self.cross_connector_var.get():
            self._refresh_cross_options()
        self._apply_cross_visibility()

        # ── 4. Rod Father ────────────────────────────────────────────────────
        rf_frame = ttk.LabelFrame(scrollable, text="Rod Father Offcuts")
        rf_frame.pack(fill="x", pady=(0, 10))

        rf_inner = ttk.Frame(rf_frame)
        rf_inner.pack(fill="x", padx=12, pady=12)

        self._load_rodfather_excel()

        # ── Two-row request table ────────────────────────────────────────────
        req_table = ttk.Frame(rf_inner)
        req_table.pack(fill="x", pady=(0, 8))

        headers = [("Side", 6), ("Material", 20), ("Type", 14), ("Required length (mm)", 18)]
        for col_i, (text, _) in enumerate(headers):
            ttk.Label(req_table, text=text, font=("Segoe UI", 10, "bold")).grid(
                row=0, column=col_i, sticky="w", padx=(0, 16), pady=(0, 4)
            )

        # Left row
        ttk.Label(req_table, text="Left", font=FONT).grid(row=1, column=0, sticky="w", padx=(0, 16), pady=(0, 4))
        ttk.Label(req_table, textvariable=self.left_rf_material_var, font=FONT, width=20, anchor="w").grid(
            row=1, column=1, sticky="w", padx=(0, 16), pady=(0, 4)
        )
        ttk.Combobox(
            req_table, textvariable=self.left_rf_shape_var,
            values=["Hex", "Cylindrical"], state="readonly", width=12
        ).grid(row=1, column=2, sticky="w", padx=(0, 16), pady=(0, 4))
        ttk.Entry(req_table, textvariable=self.left_rf_length_var, width=10).grid(
            row=1, column=3, sticky="w", padx=(0, 16), pady=(0, 4)
        )

        # Right row
        ttk.Label(req_table, text="Right", font=FONT).grid(row=2, column=0, sticky="w", padx=(0, 16))
        ttk.Label(req_table, textvariable=self.right_rf_material_var, font=FONT, width=20, anchor="w").grid(
            row=2, column=1, sticky="w", padx=(0, 16)
        )
        ttk.Combobox(
            req_table, textvariable=self.right_rf_shape_var,
            values=["Hex", "Cylindrical"], state="readonly", width=12
        ).grid(row=2, column=2, sticky="w", padx=(0, 16))
        ttk.Entry(req_table, textvariable=self.right_rf_length_var, width=10).grid(
            row=2, column=3, sticky="w", padx=(0, 16)
        )

        # Shared allocate button
        ttk.Button(
            rf_inner, text="Search and Allocate Both Sides",
            command=self._allocate_both_sides
        ).pack(anchor="w", pady=(0, 8))

        # Status lines
        self.offcuts_status_label = ttk.Label(rf_inner, text="", font=("Segoe UI", 10))
        self.offcuts_status_label.pack(anchor="w", pady=(0, 2))

        self.alloc_status_label = ttk.Label(rf_inner, text="", font=("Segoe UI", 10))
        self.alloc_status_label.pack(anchor="w", pady=(0, 6))

        # Bin overview
        bin_frame = ttk.LabelFrame(rf_inner, text="Bin Overview")
        bin_frame.pack(fill="x", pady=(0, 8))

        ttk.Label(bin_frame, text="Bin (mm)", font=("Segoe UI", 10, "bold")).grid(row=0, column=0, padx=8, sticky="w")
        ttk.Label(bin_frame, text="Total", font=("Segoe UI", 10, "bold")).grid(row=0, column=1, padx=8, sticky="e")

        for row_i, b in enumerate((100, 200, 300, 400, 600), start=1):
            ttk.Label(bin_frame, text=str(b)).grid(row=row_i, column=0, padx=8, sticky="w")
            lbl = ttk.Label(bin_frame, text="0")
            lbl.grid(row=row_i, column=1, padx=8, sticky="e")
            self.bin_labels[b] = lbl

        self._update_bin_overview()

        # Results table
        cols = ("Offcut ID", "Material", "Shape", "Length", "Status")
        self.offcuts_tree = ttk.Treeview(rf_inner, columns=cols, show="headings", height=6)
        for c in cols:
            self.offcuts_tree.heading(c, text=c)
            self.offcuts_tree.column(c, width=120, anchor="w")
        self.offcuts_tree.column("Length", width=80, anchor="e")
        self.offcuts_tree.column("Status", width=100, anchor="w")
        self.offcuts_tree.pack(fill="x", pady=(0, 8))

        # Manual reserve actions
        actions = ttk.Frame(rf_inner)
        actions.pack(fill="x")
        ttk.Button(actions, text="Reserve for Left", command=lambda: self._reserve_selected("left")).pack(side="left")
        ttk.Button(actions, text="Reserve for Right", command=lambda: self._reserve_selected("right")).pack(side="left", padx=(8, 0))
        ttk.Button(actions, text="Reload Excel", command=self._reload_rodfather).pack(side="left", padx=(8, 0))

        chosen = ttk.Frame(rf_inner)
        chosen.pack(fill="x", pady=(8, 0))
        ttk.Label(chosen, text="Chosen Left Offcut:", font=("Segoe UI", 10)).grid(row=0, column=0, sticky="w")
        ttk.Label(chosen, textvariable=self.left_offcut_id_var, font=("Segoe UI", 10, "bold")).grid(row=0, column=1, sticky="w", padx=(6, 12))
        ttk.Label(chosen, text="Chosen Right Offcut:", font=("Segoe UI", 10)).grid(row=0, column=2, sticky="w")
        ttk.Label(chosen, textvariable=self.right_offcut_id_var, font=("Segoe UI", 10, "bold")).grid(row=0, column=3, sticky="w", padx=(6, 0))

        # Seed material labels and length fields from restored state
        self._on_rod_selection_changed()
        self._sync_rf_lengths()
        self._run_offcut_query()

    # ─────────────────────────────────────────────────────────────────────────
    # Material parsing & sync
    # ─────────────────────────────────────────────────────────────────────────

    @staticmethod
    def _parse_material(rod_selection: str) -> str:
        s = rod_selection.strip()
        if s.lower().startswith("cobalt chrome"):
            return "Cobalt Chrome"
        if s.lower().startswith("titanium"):
            return "Titanium"
        return "—" if not s else s

    def _on_rod_selection_changed(self):
        self.left_rf_material_var.set(self._parse_material(self.left_rod_var.get()))
        self.right_rf_material_var.set(self._parse_material(self.right_rod_var.get()))

        # Populate predicted lengths only once both sides have been selected
        both_selected = bool(self.left_rod_var.get().strip() and self.right_rod_var.get().strip())
        if both_selected:
            if not self.left_predicted_var.get():
                self.left_predicted_var.set(str(DEFAULT_PREDICTED_LEFT_MM))
            if not self.right_predicted_var.get():
                self.right_predicted_var.set(str(DEFAULT_PREDICTED_RIGHT_MM))
        else:
            self.left_predicted_var.set("")
            self.right_predicted_var.set("")

        self._sync_rf_lengths()
        self._update_bin_overview()
        self._run_offcut_query()
        self._persist()

    def _sync_rf_lengths(self):
        left_pred = self._safe_float(self.left_predicted_var.get())
        right_pred = self._safe_float(self.right_predicted_var.get())
        self.left_rf_length_var.set(f"{left_pred:.0f}" if left_pred else "")
        self.right_rf_length_var.set(f"{right_pred:.0f}" if right_pred else "")

    # ─────────────────────────────────────────────────────────────────────────
    # Page data access
    # ─────────────────────────────────────────────────────────────────────────

    def _get_construct_levels(self):
        pd = getattr(self.app, "plan_data", {}) or {}
        ap = pd.get("anchor_planning", {})
        v = ap.get("levels", [])
        if isinstance(v, list) and v:
            return [str(x).strip() for x in v if str(x).strip()]
        return []

    def _level_index(self, lvl: str):
        if not lvl:
            return None
        s = str(lvl).strip().upper().replace(" ", "")
        m = re.match(r"^([CTLS])(\d{1,2})$", s)
        if not m:
            return None
        offsets = {"C": 0, "T": 100, "L": 200, "S": 300}
        return offsets[m.group(1)] + int(m.group(2))

    def _sort_and_dedupe_levels(self, levels):
        cleaned, seen = [], set()
        for x in levels:
            s = str(x).strip().upper().replace(" ", "")
            idx = self._level_index(s)
            if idx is None or s in seen:
                continue
            seen.add(s)
            cleaned.append((idx, s))
        cleaned.sort(key=lambda t: t[0])
        return [s for _, s in cleaned]

    # ─────────────────────────────────────────────────────────────────────────
    # Cross connector logic
    # ─────────────────────────────────────────────────────────────────────────

    def _refresh_cross_options(self):
        levels = self._sort_and_dedupe_levels(self._get_construct_levels())
        indices = [self._level_index(l) for l in levels]
        pairs = []
        for i in range(len(levels) - 1):
            ia, ib = indices[i], indices[i + 1]
            if ia is not None and ib is not None and (ib - ia) == 1:
                pairs.append(f"{levels[i]}/{levels[i+1]}")
        self.cross_options = pairs

        if self.cross_checks_frame is None:
            return

        for w in self.cross_checks_frame.winfo_children():
            w.destroy()

        self.cross_selected = [s for s in self.cross_selected if s in self.cross_options][:2]
        self.cross_check_vars = {}
        for lvl in self.cross_options:
            var = tk.BooleanVar(value=(lvl in self.cross_selected))
            self.cross_check_vars[lvl] = var
            ttk.Checkbutton(
                self.cross_checks_frame, text=lvl, variable=var,
                command=lambda L=lvl: self._on_cross_checkbox_toggled(L)
            ).pack(anchor="w")

        self._update_cross_summary()

    def _on_cross_toggle(self):
        if not self.cross_connector_var.get():
            self.cross_selected = []
            for var in getattr(self, "cross_check_vars", {}).values():
                var.set(False)
            self._update_cross_summary()
        else:
            self._refresh_cross_options()
        self._apply_cross_visibility()
        self._persist()

    def _apply_cross_visibility(self):
        if self.cross_connector_var.get():
            self.cross_section_frame.pack(fill="x", pady=(10, 0))
        else:
            self.cross_section_frame.pack_forget()

    def _on_cross_checkbox_toggled(self, toggled_level: str):
        selected = [lvl for lvl, var in self.cross_check_vars.items() if var.get()]
        if len(selected) > 2:
            self.cross_check_vars[toggled_level].set(False)
            messagebox.showwarning("Limit Reached", "You can select up to 2 cross-connector levels.")
            selected = [lvl for lvl, var in self.cross_check_vars.items() if var.get()]
        self.cross_selected = selected
        self._update_cross_summary()
        self._persist()

    def _update_cross_summary(self):
        if self.cross_summary_label is None:
            return
        self.cross_summary_label.configure(
            text="Selected: " + (", ".join(self.cross_selected) if self.cross_selected else "None")
        )

    # ─────────────────────────────────────────────────────────────────────────
    # Persist / restore
    # ─────────────────────────────────────────────────────────────────────────

    def _persist(self):
        self.app.plan_data.setdefault("rod_selection", {})
        rs = self.app.plan_data["rod_selection"]

        rs["left_rod"] = self.left_rod_var.get().strip()
        rs["right_rod"] = self.right_rod_var.get().strip()
        rs["other_rod"] = self.other_rod_var.get().strip()
        rs["cranial_cross_connector"] = bool(self.cross_connector_var.get())
        rs["cross_connector_positions"] = list(self.cross_selected)
        rs["left_offcut_id"] = self.left_offcut_id_var.get().strip()
        rs["right_offcut_id"] = self.right_offcut_id_var.get().strip()
        rs["left_predicted_length_mm"] = self._safe_float(self.left_predicted_var.get())
        rs["right_predicted_length_mm"] = self._safe_float(self.right_predicted_var.get())

        self.app.is_dirty = True

    def _restore(self):
        rs = self.app.plan_data.get("rod_selection", {})
        if not isinstance(rs, dict):
            rs = {}

        self.left_rod_var.set(str(rs.get("left_rod", "")).strip())
        self.right_rod_var.set(str(rs.get("right_rod", "")).strip())
        self.other_rod_var.set(str(rs.get("other_rod", "")).strip())
        self.cross_connector_var.set(bool(rs.get("cranial_cross_connector", False)))

        saved = rs.get("cross_connector_positions", [])
        self.cross_selected = [str(x).strip() for x in saved if str(x).strip()] if isinstance(saved, list) else []

        self.left_offcut_id_var.set(str(rs.get("left_offcut_id", "")).strip())
        self.right_offcut_id_var.set(str(rs.get("right_offcut_id", "")).strip())

        left_pred = rs.get("left_predicted_length_mm", None)
        right_pred = rs.get("right_predicted_length_mm", None)
        self.left_predicted_var.set(str(left_pred) if left_pred else "")
        self.right_predicted_var.set(str(right_pred) if right_pred else "")

    @staticmethod
    def _safe_float(s):
        try:
            v = float(str(s).strip())
            return v if v > 0 else None
        except (ValueError, TypeError):
            return None

    # ─────────────────────────────────────────────────────────────────────────
    # Navigation
    # ─────────────────────────────────────────────────────────────────────────

    def _next_to_page_12(self):
        self._persist()
        if hasattr(self.app, "setup_page_12"):
            self.app.setup_page_12()
        else:
            messagebox.showinfo("Next Page", "Page 12 not implemented yet.")

    # ─────────────────────────────────────────────────────────────────────────
    # Rod Father — Excel I/O
    # ─────────────────────────────────────────────────────────────────────────

    def _reload_rodfather(self):
        self._load_rodfather_excel()
        self._run_offcut_query()

    def _load_rodfather_excel(self):
        self.offcuts = []
        self.offcuts_load_error = None

        if not os.path.exists(self.rodfather_path):
            self.offcuts_load_error = f"Rod Father file not found: {self.rodfather_path}"
            return

        try:
            wb = openpyxl.load_workbook(self.rodfather_path)
            if "Offcuts" not in wb.sheetnames:
                self.offcuts_load_error = "Missing sheet: Offcuts"
                return

            ws = wb["Offcuts"]
            header = [str(cell.value).strip() if cell.value is not None else "" for cell in ws[1]]

            required = ["Offcut ID", "Material", "Shape", "Length", "Status"]
            missing = [c for c in required if c not in header]
            if missing:
                self.offcuts_load_error = "Missing columns in Offcuts: " + ", ".join(missing)
                return

            idx = {name: header.index(name) for name in required}

            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row:
                    continue
                offcut_id = row[idx["Offcut ID"]]
                if offcut_id is None or str(offcut_id).strip() == "":
                    continue
                try:
                    length_val = float(row[idx["Length"]])
                except Exception:
                    continue

                self.offcuts.append({
                    "offcut_id": str(offcut_id).strip(),
                    "material": str(row[idx["Material"]]).strip() if row[idx["Material"]] is not None else "",
                    "shape": str(row[idx["Shape"]]).strip() if row[idx["Shape"]] is not None else "",
                    "length": length_val,
                    "status": str(row[idx["Status"]]).strip() if row[idx["Status"]] is not None else "",
                })

        except Exception as e:
            self.offcuts_load_error = f"Failed to load Rod Father Excel: {e}"

        self._update_bin_overview()

    def _run_offcut_query(self):
        if self.offcuts_status_label is None or self.offcuts_tree is None:
            return

        if self.offcuts_load_error:
            self.offcuts_status_label.configure(text=self.offcuts_load_error)
            self._clear_offcuts_tree()
            return

        material = self.left_rf_material_var.get().strip()
        shape = self.left_rf_shape_var.get().strip()

        if material == "—":
            self._clear_offcuts_tree()
            self.offcuts_status_label.configure(text="Select a rod diameter above to filter inventory.")
            return

        results = [
            o for o in self.offcuts
            if o["material"] == material and o["shape"] == shape
            and o["status"].lower() == "available"
        ]
        results.sort(key=lambda x: x["length"])
        self._fill_offcuts_tree(results)
        self.offcuts_status_label.configure(text=f"Available: {len(results)} rods for {material}, {shape}.")

    def _clear_offcuts_tree(self):
        for item in self.offcuts_tree.get_children():
            self.offcuts_tree.delete(item)

    def _fill_offcuts_tree(self, rows):
        self._clear_offcuts_tree()
        for o in rows:
            self.offcuts_tree.insert(
                "", "end",
                values=(o["offcut_id"], o["material"], o["shape"], f"{o['length']:.0f}", o["status"])
            )

    def _get_selected_offcut_id(self):
        if self.offcuts_tree is None:
            return None
        sel = self.offcuts_tree.selection()
        if not sel:
            return None
        values = self.offcuts_tree.item(sel[0], "values")
        return str(values[0]).strip() if values else None

    def _reserve_selected(self, side: str):
        offcut_id = self._get_selected_offcut_id()
        if not offcut_id:
            messagebox.showinfo("Select an offcut", "Select an offcut row first.")
            return

        ok, err = self._set_offcut_status_in_excel(offcut_id, "Reserved")
        if not ok:
            messagebox.showerror("Reserve failed", err or "Unknown error")
            return

        for o in self.offcuts:
            if o["offcut_id"] == offcut_id:
                o["status"] = "Reserved"
                break

        self.app.plan_data.setdefault("rod_selection", {})
        rs = self.app.plan_data["rod_selection"]
        rs[f"{side}_offcut_id"] = offcut_id

        if side == "left":
            self.left_offcut_id_var.set(offcut_id)
        else:
            self.right_offcut_id_var.set(offcut_id)

        self.app.is_dirty = True
        self._update_bin_overview()
        self._run_offcut_query()

    def _set_offcut_status_in_excel(self, offcut_id: str, new_status: str):
        if not os.path.exists(self.rodfather_path):
            return False, f"Rod Father file not found: {self.rodfather_path}"

        try:
            wb = openpyxl.load_workbook(self.rodfather_path)
            ws = wb["Offcuts"]
            header = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]

            if "Offcut ID" not in header or "Status" not in header:
                return False, "Offcuts sheet missing Offcut ID or Status columns."

            id_col = header.index("Offcut ID") + 1
            st_col = header.index("Status") + 1
            found_row = None

            for r in range(2, ws.max_row + 1):
                v = ws.cell(row=r, column=id_col).value
                if v is not None and str(v).strip() == offcut_id:
                    found_row = r
                    break

            if found_row is None:
                return False, f"Offcut ID not found: {offcut_id}"

            ws.cell(row=found_row, column=st_col).value = new_status
            wb.save(self.rodfather_path)
            return True, None

        except Exception as e:
            return False, str(e)

    def _update_bin_overview(self):
        if not getattr(self, "bin_labels", None):
            return

        material = self.left_rf_material_var.get().strip()
        shape = self.left_rf_shape_var.get().strip()

        filtered = [
            o for o in self.offcuts
            if o.get("material") == material and o.get("shape") == shape
            and str(o.get("status", "")).strip().lower() == "available"
        ]

        overview = compute_bin_overview(filtered)
        for b in (100, 200, 300, 400, 600):
            lbl = self.bin_labels.get(b)
            if lbl:
                lbl.config(text=str(overview.get(b, {}).get("total", 0)))

    # ─────────────────────────────────────────────────────────────────────────
    # Allocation helpers
    # ─────────────────────────────────────────────────────────────────────────

    def _find_exact_available_piece(self, material, shape, required_len, tol=0.5):
        for o in self.offcuts:
            if o.get("material") != material or o.get("shape") != shape:
                continue
            if str(o.get("status", "")).strip().lower() != "available":
                continue
            try:
                if abs(float(o.get("length", 0.0)) - required_len) <= tol:
                    return o
            except Exception:
                continue
        return None

    def _find_best_longer_available_piece(self, material, shape, required_len):
        candidates = []
        for o in self.offcuts:
            if o.get("material") != material or o.get("shape") != shape:
                continue
            if str(o.get("status", "")).strip().lower() != "available":
                continue
            try:
                L = float(o.get("length", 0.0))
                if L >= required_len:
                    candidates.append((L, o))
            except Exception:
                continue
        if not candidates:
            return None
        candidates.sort(key=lambda t: t[0])
        return candidates[0][1]

    def _next_offcut_id(self):
        max_n = 0
        for o in self.offcuts:
            m = re.search(r"(\d+)$", str(o.get("offcut_id", "")))
            if m:
                try:
                    max_n = max(max_n, int(m.group(1)))
                except Exception:
                    pass
        return f"AUTO{max_n + 1:04d}"

    def _append_new_offcut_to_excel(self, material, shape, length_mm, status="Available"):
        if not os.path.exists(self.rodfather_path):
            return False, None, f"Rod Father file not found: {self.rodfather_path}"

        try:
            wb = openpyxl.load_workbook(self.rodfather_path)
            if "Offcuts" not in wb.sheetnames:
                return False, None, "Missing sheet: Offcuts"

            ws = wb["Offcuts"]
            header = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
            required_cols = ["Offcut ID", "Material", "Shape", "Length", "Status"]
            missing = [c for c in required_cols if c not in header]
            if missing:
                return False, None, "Missing columns: " + ", ".join(missing)

            col = {name: header.index(name) + 1 for name in required_cols}
            new_id = self._next_offcut_id()
            new_row = ws.max_row + 1

            ws.cell(row=new_row, column=col["Offcut ID"]).value = new_id
            ws.cell(row=new_row, column=col["Material"]).value = material
            ws.cell(row=new_row, column=col["Shape"]).value = shape
            ws.cell(row=new_row, column=col["Length"]).value = float(length_mm)
            ws.cell(row=new_row, column=col["Status"]).value = status

            wb.save(self.rodfather_path)
            return True, new_id, None

        except Exception as e:
            return False, None, str(e)

    def _allocate_one_side(self, side: str):
        """
        Attempt allocation for one side.
        Returns (True, msg) on success, (False, msg) on error,
        or ("needs_confirm", info_dict) when a cut requires user confirmation.
        """
        if side == "left":
            material = self.left_rf_material_var.get().strip()
            shape = self.left_rf_shape_var.get().strip()
            length_str = self.left_rf_length_var.get().strip()
        else:
            material = self.right_rf_material_var.get().strip()
            shape = self.right_rf_shape_var.get().strip()
            length_str = self.right_rf_length_var.get().strip()

        if material == "—":
            return False, f"{side.capitalize()}: no material — select a rod diameter first."
        if not length_str:
            return False, f"{side.capitalize()}: required length is empty."

        try:
            required_len = float(length_str)
            if required_len <= 0:
                raise ValueError()
        except ValueError:
            return False, f"{side.capitalize()}: invalid length '{length_str}'."

        # Exact match
        exact = self._find_exact_available_piece(material, shape, required_len)
        if exact:
            source_id = str(exact["offcut_id"]).strip()
            ok, err = self._set_offcut_status_in_excel(source_id, "Reserved")
            if not ok:
                return False, f"{side.capitalize()}: reserve failed — {err}"

            for o in self.offcuts:
                if o.get("offcut_id") == source_id:
                    o["status"] = "Reserved"
                    break

            self.app.plan_data.setdefault("rod_selection", {})
            rs = self.app.plan_data["rod_selection"]
            rs[f"{side}_offcut_id"] = source_id
            rs.setdefault("rod_father", {})[side] = {
                "used": True, "mode": "exact",
                "material": material, "type": shape,
                "required_length_mm": required_len,
                "source_offcut_id": source_id,
                "source_length_mm": float(exact.get("length", required_len)),
                "leftover_length_mm": 0.0, "leftover_offcut_id": None,
            }

            if side == "left":
                self.left_offcut_id_var.set(source_id)
            else:
                self.right_offcut_id_var.set(source_id)

            return True, f"{side.capitalize()}: exact match — reserved {source_id}."

        # Cut from longer piece
        source = self._find_best_longer_available_piece(material, shape, required_len)
        if not source:
            return False, f"{side.capitalize()}: no rod >= {required_len:.0f} mm available for {material}, {shape}."

        return "needs_confirm", {
            "side": side, "material": material, "shape": shape,
            "required_len": required_len,
            "source_id": str(source["offcut_id"]).strip(),
            "source_len": float(source.get("length", 0.0)),
            "leftover": float(source.get("length", 0.0)) - required_len,
        }

    def _confirm_and_cut(self, info: dict):
        side, material, shape = info["side"], info["material"], info["shape"]
        required_len, source_id = info["required_len"], info["source_id"]
        source_len, leftover = info["source_len"], info["leftover"]

        ok, err = self._set_offcut_status_in_excel(source_id, "Reserved")
        if not ok:
            return False, f"{side.capitalize()}: reserve failed — {err}"

        for o in self.offcuts:
            if o.get("offcut_id") == source_id:
                o["status"] = "Reserved"
                break

        new_offcut_id = None
        if leftover > 0.5:
            ok2, new_id, err2 = self._append_new_offcut_to_excel(material, shape, leftover)
            if not ok2:
                return False, f"{side.capitalize()}: reserved source but failed to add leftover — {err2}"
            new_offcut_id = new_id
            self.offcuts.append({
                "offcut_id": str(new_id).strip(),
                "material": material, "shape": shape,
                "length": float(leftover), "status": "Available",
            })

        self.app.plan_data.setdefault("rod_selection", {})
        rs = self.app.plan_data["rod_selection"]
        rs[f"{side}_offcut_id"] = source_id
        rs.setdefault("rod_father", {})[side] = {
            "used": True, "mode": "cut",
            "material": material, "type": shape,
            "required_length_mm": required_len,
            "source_offcut_id": source_id,
            "source_length_mm": source_len,
            "leftover_length_mm": leftover,
            "leftover_offcut_id": new_offcut_id,
        }

        if side == "left":
            self.left_offcut_id_var.set(source_id)
        else:
            self.right_offcut_id_var.set(source_id)

        leftover_note = (
            f", created {leftover:.0f} mm offcut ({new_offcut_id})"
            if new_offcut_id else ", no leftover recorded"
        )
        return True, (
            f"{side.capitalize()}: cut {required_len:.0f} mm from "
            f"{source_len:.0f} mm ({source_id}){leftover_note}."
        )

    def _allocate_both_sides(self):
        results = {}
        confirms = {}

        for side in ("left", "right"):
            outcome = self._allocate_one_side(side)
            if outcome[0] == "needs_confirm":
                confirms[side] = outcome[1]
            else:
                results[side] = outcome  # (True/False, message)

        # Single combined confirmation dialog for any cuts needed
        if confirms:
            lines = ["No exact match found for the following side(s):\n"]
            for side, info in confirms.items():
                leftover_note = (
                    f"{info['leftover']:.0f} mm leftover offcut will be created"
                    if info["leftover"] > 0.5 else "no leftover (too small)"
                )
                lines.append(
                    f"  {side.capitalize()}: cut {info['required_len']:.0f} mm from "
                    f"{info['source_len']:.0f} mm (Offcut ID {info['source_id']}) — {leftover_note}"
                )
            lines.append("\nDo you want to proceed with these cuts?")

            if messagebox.askyesno("Confirm Cuts", "\n".join(lines)):
                for side, info in confirms.items():
                    results[side] = self._confirm_and_cut(info)
            else:
                for side in confirms:
                    results[side] = (False, f"{side.capitalize()}: cancelled.")

        self.app.is_dirty = True
        self._update_bin_overview()
        self._run_offcut_query()

        messages = [results[s][1] for s in ("left", "right") if s in results]
        if messages:
            self.alloc_status_label.config(text="  |  ".join(messages))